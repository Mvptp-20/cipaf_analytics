"""
CIPAF - Research Triage Tool
==============================
Scans all downloaded PDF and Excel files in a folder,
extracts key text, and produces _INVENTARIO_<TOPIC>.xlsx

Usage (called by Claude via the pipeline skill):
    python triage_cipaf.py \
        --folder "/path/to/research/folder" \
        --topic-slug "pobreza" \
        --keywords-json "/path/to/keywords.json"  # optional

If --keywords-json is not provided, uses the default poverty keyword set.
"""

import os
import re
import json
import argparse
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# ARGUMENT PARSING
# ─────────────────────────────────────────────

parser = argparse.ArgumentParser(description="CIPAF Research Triage Tool")
parser.add_argument("--folder",        required=True, help="Research folder to scan")
parser.add_argument("--topic-slug",    default="inventario", help="Topic slug for output filename")
parser.add_argument("--keywords-json", help="Path to JSON file with custom keywords by category")
args = parser.parse_args()

RESEARCH_FOLDER = args.folder
TOPIC_SLUG      = args.topic_slug
OUTPUT_FILE     = os.path.join(RESEARCH_FOLDER, f"_INVENTARIO_{TOPIC_SLUG.upper()}.xlsx")

# ─────────────────────────────────────────────
# KEYWORDS
# ─────────────────────────────────────────────

DEFAULT_KEYWORDS = {
    "Pobreza por Provincia/Región": [
        "provincia", "provincial", "región", "regional", "macrorregión",
        "distrito", "municipio", "geográfico", "territorio", "zona",
        "desagregación geográfica", "brecha regional", "disparidad territorial",
    ],
    "Género / Mujeres": [
        "mujer", "mujeres", "género", "femenino", "femenina",
        "jefa de hogar", "jefatura femenina", "brecha de género",
        "feminización", "violencia de género", "desigualdad de género",
        "brechas salariales", "trabajo doméstico", "economía del cuidado",
        "empoderamiento", "autonomía económica",
    ],
    "Pobreza Multidimensional": [
        "multidimensional", "ipm", "privación", "privaciones",
        "necesidades básicas insatisfechas", "nbi", "carencia", "carencias",
        "índice de pobreza multidimensional", "mpi", "alkire", "foster",
        "bienestar", "calidad de vida", "desarrollo humano", "idh",
    ],
    "Niñez e Infancia": [
        "niñez", "infancia", "niño", "niña", "niños", "niñas",
        "primera infancia", "adolescente", "adolescencia",
        "pobreza infantil", "desarrollo infantil", "nutrición infantil",
        "trabajo infantil", "deserción escolar", "mortalidad infantil", "unicef",
    ],
    "Mercado Laboral / Empleo": [
        "empleo", "desempleo", "desocupación", "ocupación", "trabajo",
        "informalidad", "empleo informal", "precariedad laboral",
        "subempleo", "ingreso laboral", "salario", "remuneración",
        "mercado de trabajo", "tasa de actividad", "pea", "cuenta propia",
    ],
    "Políticas Sociales / Transferencias": [
        "política social", "políticas sociales", "transferencia", "transferencias",
        "plan social", "programa social", "subsidio", "beneficiario",
        "protección social", "seguridad social", "pensión", "jubilación",
        "intervención social", "focalización",
    ],
    "Alimentación / Seguridad Alimentaria": [
        "alimentación", "seguridad alimentaria", "inseguridad alimentaria",
        "hambre", "malnutrición", "desnutrición", "nutrición",
        "canasta básica alimentaria", "cba", "canasta básica total",
        "pobreza alimentaria", "indigencia",
    ],
    "Vivienda / Hábitat": [
        "vivienda", "hábitat", "hacinamiento", "déficit habitacional",
        "agua potable", "saneamiento", "asentamiento", "villa",
        "barrio informal", "precariedad habitacional",
    ],
    "Educación": [
        "educación", "escolarización", "acceso a la educación",
        "abandono escolar", "analfabetismo", "alfabetización",
        "nivel educativo", "brecha educativa", "rezago escolar",
    ],
    "Salud": [
        "salud", "acceso a la salud", "sistema de salud", "cobertura de salud",
        "mortalidad", "morbilidad", "esperanza de vida", "salud materno-infantil",
        "salud reproductiva", "discapacidad", "salud mental",
    ],
    "Metodología / Medición": [
        "metodología", "medición", "encuesta", "muestra", "muestreo",
        "eph", "encuesta permanente de hogares", "cepal", "banco mundial",
        "línea de pobreza", "umbral de pobreza", "índice", "indicador",
        "estadística", "microdatos", "panel de datos",
    ],
    "Adultos Mayores": [
        "adulto mayor", "adultos mayores", "vejez", "envejecimiento",
        "persona mayor", "tercera edad", "jubilado", "pensionado",
        "pobreza en la vejez", "sistema previsional",
    ],
    "Pueblos Indígenas / Comunidades": [
        "indígena", "indígenas", "pueblo originario", "pueblos originarios",
        "comunidad", "comunidades", "interculturalidad", "etnia",
        "pobreza indígena", "mapuche", "quechua", "guaraní", "wichí", "qom",
    ],
}

if args.keywords_json:
    with open(args.keywords_json, encoding="utf-8") as f:
        KEYWORDS = json.load(f)
else:
    KEYWORDS = DEFAULT_KEYWORDS

PDF_PREVIEW_CHARS = 5000

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def get_subfolders(base: str):
    folders = [base]
    for item in Path(base).iterdir():
        if item.is_dir() and not item.name.startswith("_") and item.name not in ("HIGH","MEDIUM","LOW","UNCATALOGUED"):
            folders.append(str(item))
    return folders


def extract_pdf_text(filepath: str, max_chars: int = PDF_PREVIEW_CHARS) -> str:
    try:
        import fitz
        doc  = fitz.open(filepath)
        text = ""
        for page in doc:
            text += page.get_text()
            if len(text) >= max_chars:
                break
        doc.close()
        return text[:max_chars].lower()
    except Exception as e:
        return f"[ERROR reading PDF: {e}]"


def extract_excel_text(filepath: str) -> str:
    try:
        wb    = load_workbook(filepath, read_only=True, data_only=True)
        parts = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        parts.append(cell)
        wb.close()
        return " ".join(parts).lower()
    except Exception as e:
        return f"[ERROR reading Excel: {e}]"


def extract_title_from_pdf(filepath: str) -> str:
    try:
        import fitz
        doc        = fitz.open(filepath)
        page       = doc[0]
        blocks     = page.get_text("blocks")
        doc.close()
        candidates = []
        for b in blocks:
            if b[6] == 0:
                text = b[4].strip().replace("\n", " ")
                if len(text) > 10:
                    candidates.append(text)
        return candidates[0][:120] if candidates else ""
    except Exception:
        return ""


def score_file(text: str) -> dict:
    scores    = {}
    total_hits = 0
    for category, keywords in KEYWORDS.items():
        hits = sum(1 for kw in keywords if kw in text)
        scores[category] = hits
        total_hits += hits
    categories_hit = sum(1 for v in scores.values() if v > 0)
    if categories_hit >= 4:
        relevance = "🟢 HIGH"
    elif categories_hit >= 2:
        relevance = "🟡 MEDIUM"
    else:
        relevance = "🔴 LOW"
    return {**scores, "Relevance": relevance, "Total Hits": total_hits}


def size_label(filepath: str) -> str:
    kb = os.path.getsize(filepath) / 1024
    return f"{kb/1024:.1f} MB" if kb >= 1024 else f"{kb:.0f} KB"


# ─────────────────────────────────────────────
# MAIN SCAN
# ─────────────────────────────────────────────

def scan_files():
    rows    = []
    folders = get_subfolders(RESEARCH_FOLDER)
    COLUMNS_CAT = list(KEYWORDS.keys())

    for folder in folders:
        source = os.path.basename(folder) if folder != RESEARCH_FOLDER else "ROOT"
        for fname in sorted(os.listdir(folder)):
            fpath = os.path.join(folder, fname)
            ext   = Path(fname).suffix.lower()
            if ext not in (".pdf", ".xlsx", ".xls", ".csv"):
                continue
            if fname.startswith("_"):
                continue

            print(f"  Scanning: {fname}")

            if ext == ".pdf":
                text      = extract_pdf_text(fpath)
                doc_title = extract_title_from_pdf(fpath)
                file_type = "PDF"
            elif ext in (".xlsx", ".xls"):
                text      = extract_excel_text(fpath)
                doc_title = ""
                file_type = "Excel"
            elif ext == ".csv":
                try:
                    df        = pd.read_csv(fpath, nrows=50, encoding="latin-1", on_bad_lines='skip')
                    text      = " ".join(df.columns.tolist()).lower()
                    doc_title = ""
                    file_type = "CSV"
                except Exception:
                    text = ""; doc_title = ""; file_type = "CSV"
            else:
                continue

            scores = score_file(text)

            row = {
                "Archivo":          fname,
                "Título Detectado": doc_title if doc_title else fname,
                "Tipo":             file_type,
                "Fuente":           source,
                "Tamaño":           size_label(fpath),
                "Relevancia":       scores["Relevance"],
            }
            for cat in COLUMNS_CAT:
                short = cat.split("/")[0].strip()[:20]
                row[short] = "✔" if scores.get(cat, 0) > 0 else ""

            row["Total Coincidencias"] = scores["Total Hits"]
            row["Notas"] = ""
            rows.append(row)

    return sorted(rows, key=lambda x: (x["Relevancia"], -x["Total Coincidencias"]))


# ─────────────────────────────────────────────
# BUILD EXCEL OUTPUT
# ─────────────────────────────────────────────

COLOR_HEADER_BG   = "3B092B"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_HIGH        = "F2D4EC"
COLOR_MED         = "FFF9C4"
COLOR_LOW         = "F5F5F5"
COLOR_CHECK       = "BA1C88"


def build_excel(rows: list):
    if not rows:
        print("No files found to catalog.")
        return

    df = pd.DataFrame(rows)
    df.to_excel(OUTPUT_FILE, index=False, sheet_name="Inventario")

    wb = load_workbook(OUTPUT_FILE)
    ws = wb["Inventario"]
    thin   = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cols   = [c.value for c in ws[1]]

    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = Font(name="Arial", bold=True, color=COLOR_HEADER_FONT, size=10)
        cell.fill      = PatternFill("solid", start_color=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.row_dimensions[1].height = 32

    for row_idx, row_data in enumerate(rows, start=2):
        relevance = row_data.get("Relevancia", "")
        row_color = COLOR_HIGH if "HIGH" in relevance else (COLOR_MED if "MEDIUM" in relevance else COLOR_LOW)
        for col_idx in range(1, len(cols) + 1):
            cell      = ws.cell(row=row_idx, column=col_idx)
            is_check  = cell.value == "✔"
            cell.font = Font(name="Arial", size=9, color=COLOR_CHECK if is_check else "000000", bold=is_check)
            cell.fill = PatternFill("solid", start_color=row_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ws2  = wb.create_sheet("Resumen")
    high = sum(1 for r in rows if "HIGH"   in r.get("Relevancia",""))
    med  = sum(1 for r in rows if "MEDIUM" in r.get("Relevancia",""))
    low  = sum(1 for r in rows if "LOW"    in r.get("Relevancia",""))

    summary = [
        ["CIPAF — Inventario de Investigación", ""],
        ["", ""],
        ["TOTAL ARCHIVOS ESCANEADOS", len(rows)],
        ["", ""],
        ["RELEVANCIA", ""],
        ["🟢 Alta relevancia (4+ categorías)",  high],
        ["🟡 Media relevancia (2-3 categorías)", med],
        ["🔴 Baja relevancia (0-1 categorías)",  low],
    ]
    for r_idx, (label, value) in enumerate(summary, start=1):
        ws2.cell(row=r_idx, column=1).value = label
        ws2.cell(row=r_idx, column=2).value = value
        if r_idx == 1:
            ws2.cell(row=r_idx, column=1).font = Font(name="Arial", bold=True, size=13, color=COLOR_HEADER_BG)

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 14

    wb.save(OUTPUT_FILE)
    print(f"\n✔  Inventario guardado en: {OUTPUT_FILE}")
    print(f"   🟢 Alta    : {high}  |  🟡 Media : {med}  |  🔴 Baja : {low}")
    print(f"TRIAGE_OUTPUT:{OUTPUT_FILE}")


if __name__ == "__main__":
    print("CIPAF Triage Tool — Starting scan...\n")
    rows = scan_files()
    print(f"\nBuilding Excel inventory ({len(rows)} files)...")
    build_excel(rows)
    print("\nDone.")